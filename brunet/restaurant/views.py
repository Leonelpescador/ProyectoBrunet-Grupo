from .forms import ModificarPedidoForm  
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Caja, Pedido, Reserva, Inventario, Proveedor, Compra, Mesa, Pago, TransaccionCaja
from .forms import (
    AperturaCajaForm, 
    CierreCajaForm, 
    PedidoForm, 
    ModificarPedidoForm, 
    PagoForm, 
    ModificarPagoForm, 
    ReservaForm, 
    ModificarReservaForm, 
    CompraForm, 
    ModificarCompraForm, 
    MesaForm, 
    ProveedorForm, 
    InventarioForm,
    DetalleCompraForm
    
)
from django.contrib.auth.models import User

# Home View
@login_required
def home(request):
    caja_abierta = Caja.objects.filter(usuario=request.user, estado='abierta').first()
    
    mesas = Mesa.objects.all()

    context = {
        'mesas': mesas,
        'caja_abierta': caja_abierta,
    }

    return render(request, 'home.html', context)



# Creación de Pedido aca comienza
from .models import DetallePedido
from .forms import PedidoForm, DetallePedidoForm
from django.forms import inlineformset_factory

from django.forms import inlineformset_factory
from .models import Pedido, DetallePedido

# Formset para DetallePedido
DetallePedidoFormSet = inlineformset_factory(Pedido, DetallePedido, form=DetallePedidoForm, extra=1)



from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Pedido, DetallePedido, Mesa, Menu
import json

def crear_pedido(request, mesa_id):
    mesa = get_object_or_404(Mesa, id=mesa_id)
    categorias = Categoria.objects.all()  # Filtrar por categorías
    platos = Menu.objects.filter(disponible=True)  # Todos los platos disponibles

    if request.method == 'POST':
        # Procesar el pedido enviado por el usuario
        pedido_data = json.loads(request.body.decode('utf-8'))
        
        # Crear el pedido
        pedido = Pedido.objects.create(mesa=mesa, usuario=request.user, estado='preparando', total=0)
        
        # Agregar detalles del pedido
        for detalle in pedido_data['platos']:
            plato = get_object_or_404(Menu, id=detalle['plato_id'])
            cantidad = detalle['cantidad']
            DetallePedido.objects.create(
                pedido=pedido, 
                menu=plato, 
                cantidad=cantidad, 
                precio_unitario=plato.precio, 
                subtotal=plato.precio * cantidad
            )
        
        # Calcular el total del pedido
        pedido.total = sum(d.subtotal for d in pedido.detalles.all())
        pedido.save()

        # Devolver el número de comanda
        return JsonResponse({'success': True, 'numero_comanda': pedido.numero_comanda})

    return render(request, 'pedido/crear_pedido.html', {
        'mesa': mesa,
        'categorias': categorias,
        'platos': platos,
    })



@login_required
def modificar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    mesa = pedido.mesa  
    categorias = Categoria.objects.all()
    platos = Menu.objects.all()

    if request.method == 'POST':
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # Verificar si es una solicitud AJAX
            data = json.loads(request.body)
            nuevos_detalles = data.get('detalles', [])

            # Actualizar o crear nuevos detalles
            for detalle in nuevos_detalles:
                menu_id = detalle.get('menu_id')
                cantidad = detalle.get('cantidad')
                plato = Menu.objects.get(id=menu_id)

                detalle_pedido, created = DetallePedido.objects.get_or_create(
                    pedido=pedido,
                    menu=plato,
                    defaults={'cantidad': cantidad, 'precio_unitario': plato.precio}
                )
                if not created:
                    detalle_pedido.cantidad = cantidad
                    detalle_pedido.save()
            ids_nuevos_detalles = [detalle['menu_id'] for detalle in nuevos_detalles]
            pedido.detalles.exclude(menu_id__in=ids_nuevos_detalles).delete()

            return JsonResponse({'success': True, 'message': 'Pedido actualizado exitosamente.'})

        form = ModificarPedidoForm(request.POST, instance=pedido)
        if form.is_valid():
            form.save()
            return redirect('pedidos_activos')

    else:
        form = ModificarPedidoForm(instance=pedido)

    return render(request, 'pedido/modificar_pedido.html', {
        'pedido': pedido,
        'mesa': mesa,
        'categorias': categorias,
        'platos': platos,
        'form': form
    })



@login_required
def eliminar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        pedido.estado = 'eliminado'  # Cambia el estado a 'eliminado' en lugar de borrarlo
        pedido.save()
        messages.success(request, 'Pedido marcado como eliminado con éxito.')
        return redirect('pedidos_activos')

    return render(request, 'pedido/eliminar_pedido.html', {'pedido': pedido})


@login_required
def pedidos_activos(request):
    Pedido.objects.filter(en_proceso=True, usuario_procesando=request.user).update(en_proceso=False, usuario_procesando=None)
    pedidos = Pedido.objects.filter(estado__in=['pendiente', 'preparando', 'servido'])
    return render(request, 'pedido/pedidos_activos.html', {'pedidos': pedidos})


from .models import Pedido, DetallePedido
from django.core.serializers import serialize

def nuevos_pedidos(request):
    # Obtener pedidos en estado 'preparando'
    pedidos = Pedido.objects.filter(estado='preparando').order_by('fecha_pedido')

    # Formatear los pedidos para enviarlos como JSON
    pedidos_data = []
    for pedido in pedidos:
        detalles = DetallePedido.objects.filter(pedido=pedido)
        pedidos_data.append({
            'id': pedido.id,
            'numero_comanda': pedido.numero_comanda,
            'mesa': {
                'numero_mesa': pedido.mesa.numero_mesa,
            },
            'usuario': {
                'username': pedido.usuario.username,
            },
            'tiempo': (timezone.now() - pedido.fecha_pedido).total_seconds() / 60,  # Tiempo en minutos
            'detalles': [
                {
                    'menu': {
                        'nombre_plato': detalle.menu.nombre_plato
                    },
                    'cantidad': detalle.cantidad
                } for detalle in detalles
            ]
        })

    return JsonResponse({'nuevos_pedidos': pedidos_data})

def pedidos_eliminados(request):
    pedidos_eliminados_ids = Pedido.objects.filter(estado='eliminado').values_list('id', flat=True)
    return JsonResponse({'pedidos_eliminados': list(pedidos_eliminados_ids)})









from .models import Menu

def obtener_precio_plato(request):
    menu_id = request.GET.get('menu_id')
    if menu_id:
        try:
            plato = Menu.objects.get(id=menu_id)
            return JsonResponse({'precio': plato.precio})
        except Menu.DoesNotExist:
            return JsonResponse({'error': 'Plato no encontrado'}, status=404)
    return JsonResponse({'error': 'Solicitud inválida'}, status=400)


#fin pedidos

# Creación de Pago
@login_required
def crear_pago(request, pedido_id):
    # Obtener el pedido
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        form = PagoForm(request.POST)
        if form.is_valid():
            # Crear el pago y asociarlo con el pedido
            pago = form.save(commit=False)
            pago.pedido = pedido
            pago.save()

            # Imprimir el estado actual del pedido antes de cambiarlo
            print(f"Estado actual del pedido {pedido.id}: {pedido.estado}")

            # Actualizar el estado del pedido a 'pagado'
            pedido.estado = 'pagado'
            pedido.save()

            # Imprimir el estado actualizado del pedido
            print(f"Nuevo estado del pedido {pedido.id}: {pedido.estado}")

            # Mostrar mensaje de éxito
            messages.success(request, f'Pago para el pedido {pedido.id} registrado con éxito.')

            # Redirigir a la página de pedidos activos
            return redirect('pedidos_activos')

    else:
        form = PagoForm()

    return render(request, 'pago/crear_pago.html', {'form': form, 'pedido': pedido})



# Modificación de Pago
@login_required
def modificar_pago(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    if request.method == 'POST':
        form = ModificarPagoForm(request.POST, instance=pago)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pago modificado con éxito.')
            return redirect('home')
    else:
        form = ModificarPagoForm(instance=pago)
    return render(request, 'pago/modificar_pago.html', {'form': form})

# Eliminación de Pago
@login_required
def eliminar_pago(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    if request.method == 'POST':
        pago.delete()
        messages.success(request, 'Pago eliminado con éxito.')
        return redirect('home')
    return render(request, 'pago/eliminar_pago.html', {'pago': pago})




# Vista de Reservas
@login_required
def reservas(request):
    reservas = Reserva.objects.all()
    return render(request, 'reserva/reservas.html', {'reservas': reservas})

# Creación de Reserva


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .forms import ReservaForm, ModificarReservaForm
from .models import Reserva

# Creación de Reserva
@login_required
def crear_reserva(request):
    if request.method == 'POST':
        form = ReservaForm(request.POST)
        if form.is_valid():
            reserva = form.save(commit=False)  # No guardar aún en la base de datos
            reserva.usuario = request.user  # Asignar el usuario actual
            reserva.save()  # Ahora guardar en la base de datos
            messages.success(request, 'Reserva creada con éxito.')
            return redirect('reservas')  # Redirige a la lista de reservas o donde sea necesario
    else:
        form = ReservaForm()
    return render(request, 'reserva/crear_reserva.html', {'form': form})

#modifcar reserva
@login_required
def modificar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    if request.method == 'POST':
        form = ModificarReservaForm(request.POST, instance=reserva)
        if form.is_valid():
            form.save()
            messages.success(request, 'Reserva modificada con éxito.')
            return redirect('reservas')
    else:
        form = ModificarReservaForm(instance=reserva)
    return render(request, 'reserva/editar_reserva.html', {'form': form})


# Eliminación de Reserva
@login_required
def eliminar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    if request.method == 'POST':
        reserva.delete()
        messages.success(request, 'Reserva eliminada con éxito.')
        return redirect('reservas')  # Redirige a la lista de reservas o donde sea necesario
    return render(request, 'reserva/eliminar_reserva.html', {'reserva': reserva})





#fin de flujo de reserva.






# Inventario
@login_required
def inventario(request):
    inventarios = Inventario.objects.all()
    return render(request, 'inventario/inventario.html', {'inventarios': inventarios})

# Creación de Inventario
def crear_inventario(request):
    if request.method == 'POST':
        form = InventarioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inventario')
    else:
        form = InventarioForm()
    return render(request, 'inventario/crear_inventario.html', {'form': form})

# Edición de Inventario
def editar_inventario(request, pk):
    inventario = get_object_or_404(Inventario, pk=pk)
    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            form.save()
            return redirect('inventario')
    else:
        form = InventarioForm(instance=inventario)
    return render(request, 'inventario/editar_inventario.html', {'form': form})

# Eliminación de Inventario
def eliminar_inventario(request, pk):
    inventario = get_object_or_404(Inventario, pk=pk)
    if request.method == 'POST':
        inventario.delete()
        return redirect('inventario')
    return render(request, 'inventario/eliminar_inventario.html', {'inventario': inventario})


# Proveedores
@login_required
def proveedores(request):
    proveedores = Proveedor.objects.all()
    return render(request, 'proveedores/proveedores.html', {'proveedores': proveedores})

# Creación de Proveedor
@login_required
def crear_proveedor(request):
    if request.method == "POST":
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor creado exitosamente.')
            return redirect('proveedores')
        else:
            messages.error(request, 'Hubo un error al crear el proveedor. Por favor, verifica los datos ingresados.')
    else:
        form = ProveedorForm()

    return render(request, 'proveedores/crear_proveedores.html', {'form': form})  # Aquí debe coincidir con el nombre del archivo

# Edición de Proveedor
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from .forms import ProveedorForm

@login_required
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == "POST":
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.has_changed():  # Verifica si se ha realizado algún cambio
            if form.is_valid():
                form.save()
                messages.success(request, 'Proveedor editado exitosamente.')
                return redirect('proveedores')
            else:
                messages.error(request, 'Hubo un error al editar el proveedor. Por favor, verifica los datos ingresados.')
        else:
            messages.info(request, 'No se han realizado cambios.')
            return redirect('proveedores')
    else:
        form = ProveedorForm(instance=proveedor)

    return render(request, 'proveedores/editar_proveedores.html', {'form': form})




# Eliminación de Proveedor
@login_required
def eliminar_proveedor(request, pk):
    proveedor = Proveedor.objects.get(pk=pk)
    if request.method == "POST":
        proveedor.delete()
        return redirect('proveedores')
    
    return render(request, 'proveedores/eliminar_proveedores.html', {'proveedor': proveedor})  # Aquí debe coincidir con el nombre del archivo


# Compras


from django.forms import inlineformset_factory
from .models import Compra, DetalleCompra
from .forms import CompraForm, DetalleCompraForm
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required

from django.forms import inlineformset_factory
from .models import Compra, DetalleCompra, Inventario
from .forms import CompraForm, DetalleCompraForm
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required

@login_required
def crear_compra(request):
    DetalleCompraFormSet = inlineformset_factory(Compra, DetalleCompra, form=DetalleCompraForm, extra=1, can_delete=True)

    if request.method == 'POST':
        form = CompraForm(request.POST, request.FILES)
        detalle_formset = DetalleCompraFormSet(request.POST)  

        if form.is_valid() and detalle_formset.is_valid():
            compra = form.save()  
            detalle_formset.instance = compra  
            detalle_formset.save()

            # Actualizar el inventario
            for detalle in detalle_formset.cleaned_data:
                inventario_item = detalle['inventario']
                cantidad_comprada = detalle['cantidad']
                
                # Aumentar la cantidad en el inventario
                inventario_item.cantidad_actual += cantidad_comprada
                inventario_item.save()

            return redirect('compras')
    else:
        form = CompraForm()
        detalle_formset = DetalleCompraFormSet()

    return render(request, 'compra/crear_compra.html', {
        'form': form,
        'detalle_formset': detalle_formset
    })


@login_required
def compras(request):
    lista_compras = Compra.objects.all()
    return render(request, 'compra/compras.html', {'compras': lista_compras})



@login_required
def editar_compra(request, pk):
    compra = get_object_or_404(Compra, pk=pk)

    # Crear el formset para manejar DetalleCompra asociado a esta compra
    DetalleCompraFormSet = inlineformset_factory(Compra, DetalleCompra, form=DetalleCompraForm, extra=1, can_delete=True)

    if request.method == 'POST':
        form = CompraForm(request.POST, request.FILES, instance=compra)
        detalle_formset = DetalleCompraFormSet(request.POST, instance=compra)

        if form.is_valid() and detalle_formset.is_valid():
            form.save()  # Guardamos la compra editada
            detalle_formset.save()  # Guardamos los detalles de la compra
            return redirect('compras')
    else:
        form = CompraForm(instance=compra)
        detalle_formset = DetalleCompraFormSet(instance=compra)

    return render(request, 'compra/editar_compra.html', {
        'form': form,
        'detalle_formset': detalle_formset
    })
    
@login_required
def eliminar_compra(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        compra.delete()
        return redirect('compras')
    return render(request, 'compra/eliminar_compra.html', {'compra': compra})
# Mesas
@login_required
def lista_mesas(request):
    mesas = Mesa.objects.all()
    return render(request, 'mesa/mesas.html', {'mesas': mesas})
@login_required
def crear_mesa(request):
    if request.method == 'POST':
        form = MesaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_mesas')
    else:
        form = MesaForm()
    return render(request, 'mesa/crear_mesa.html', {'form': form})
@login_required
def editar_mesa(request, pk):  
    mesa = get_object_or_404(Mesa, id=pk)
    if request.method == 'POST':
        form = MesaForm(request.POST, instance=mesa)
        if form.is_valid():
            form.save()
            return redirect('lista_mesas')
    else:
        form = MesaForm(instance=mesa)
    return render(request, 'mesa/editar_mesa.html', {'form': form})
@login_required
def eliminar_mesa(request, mesa_id):
    mesa = get_object_or_404(Mesa, id=mesa_id)
    mesa.delete()
    return redirect('lista_mesas')



# Flujo de Caja
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum
from .models import Caja, Pago, Pedido, TransaccionCaja
from django.contrib.auth.decorators import login_required

# Apertura de Caja
@login_required
def apertura_caja(request):
    caja_abierta = Caja.objects.filter(usuario=request.user, estado='abierta').first()
    if caja_abierta:
        messages.error(request, 'Ya tienes una caja abierta.')
        return redirect('consulta_caja', caja_id=caja_abierta.id)

    if request.method == 'POST':
        form = AperturaCajaForm(request.POST)
        if form.is_valid():
            caja = form.save(commit=False)
            caja.usuario = request.user  
            caja.save()
            messages.success(request, 'Caja abierta con éxito.')
            return redirect('consulta_caja', caja_id=caja.id)
    else:
        form = AperturaCajaForm()
    return render(request, 'caja/apertura_caja.html', {'form': form})



# Consulta de Caja
@login_required
def consulta_caja(request, caja_id):
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener las transacciones de la caja
    transacciones = caja.transacciones.all()

    # Calcular el monto total de ingresos
    total_ingresos = transacciones.filter(tipo='ingreso').aggregate(Sum('monto'))['monto__sum'] or 0
    
    # Calcular los totales por método de pago
    total_efectivo = sum(t.monto for t in transacciones if t.pago and t.pago.metodo_pago == 'efectivo')
    total_tarjeta = sum(t.monto for t in transacciones if t.pago and t.pago.metodo_pago == 'tarjeta')
    total_transferencia = sum(t.monto for t in transacciones if t.pago and t.pago.metodo_pago == 'transferencia')

    # Obtener los pagos asociados a los pedidos de la caja
    pedidos_en_caja = Pedido.objects.filter(fecha_pedido__gte=caja.apertura, fecha_pedido__lte=caja.cierre if caja.cierre else timezone.now())
    pagos = Pago.objects.filter(pedido__in=pedidos_en_caja)

    return render(request, 'caja/consulta_caja.html', {
        'caja': caja,
        'transacciones': transacciones,
        'total_ingresos': total_ingresos,
        'pagos': pagos,
        'total_efectivo': total_efectivo,
        'total_tarjeta': total_tarjeta,
        'total_transferencia': total_transferencia,
    })




# Cierre de Caja
@login_required
def cierre_caja(request, caja_id):
    caja = get_object_or_404(Caja, id=caja_id, usuario=request.user)  # Solo cerrar su propia caja

    if caja.estado != 'abierta':  # Verificar que la caja esté abierta
        messages.error(request, 'La caja ya ha sido cerrada.')
        return redirect('cierre_caja', caja_id=caja.id)

    if request.method == 'POST':
        total_final = caja.calcular_total_final()
        caja.cerrar_caja(total_final)
        messages.success(request, f'Caja {caja.id} cerrada exitosamente con un total de {caja.total_final}.')
        
        # Después de cerrar, quedarse en la página de cierre para descargar el reporte
        total_ingresos = sum(transaccion.monto for transaccion in caja.transacciones.filter(tipo='ingreso'))
        return render(request, 'caja/cierre_caja.html', {
            'caja': caja,
            'total_final': total_final,
            'total_ingresos': total_ingresos,
            'cierre_exitoso': True,  # Añadir un indicador de cierre exitoso
        })

    # Calcular el total de ingresos de la caja
    total_final = caja.calcular_total_final()
    total_ingresos = sum(transaccion.monto for transaccion in caja.transacciones.filter(tipo='ingreso'))

    return render(request, 'caja/cierre_caja.html', {
        'caja': caja,
        'total_final': total_final,
        'total_ingresos': total_ingresos,
    })










@login_required
def marcar_servido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id, estado='preparando')

    if pedido.en_proceso and pedido.usuario_procesando != request.user:
        messages.error(request, f"Este pedido está siendo procesado por {pedido.usuario_procesando.username}.")
        return redirect('pedidos_activos')

    # Marcar el pedido como servido
    pedido.estado = 'servido'
    pedido.marcar_completado()  # Liberar el proceso si estaba marcado en proceso
    pedido.save()

    return JsonResponse({'success': True, 'message': f"El pedido {pedido.id} ha sido marcado como servido."})
#fin





#reporte de cierre de caja 
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import io
from django.utils import timezone

def descargar_reporte_caja(request, caja_id):
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Calcular los totales por método de pago
    transacciones = caja.transacciones.all()
    total_efectivo = sum(t.monto for t in transacciones if t.pago.metodo_pago == 'efectivo')
    total_tarjeta = sum(t.monto for t in transacciones if t.pago.metodo_pago == 'tarjeta')
    total_transferencia = sum(t.monto for t in transacciones if t.pago.metodo_pago == 'transferencia')
    total_final = sum(t.monto for t in transacciones)
    
    # Prepare the context
    context = {
        'caja': caja,
        'total_efectivo': total_efectivo,
        'total_tarjeta': total_tarjeta,
        'total_transferencia': total_transferencia,
        'total_final': total_final,
    }

    template_path = 'caja/reporte_caja_pdf.html'
    
    # Verificar si la caja tiene una fecha de cierre
    if caja.cierre:
        fecha_cierre = caja.cierre.strftime("%Y-%m-%d_%H-%M-%S")
    else:
        # Si la caja no está cerrada, usar la fecha y hora actual
        fecha_cierre = timezone.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    nombre_archivo = f"Cierre_de_caja_{fecha_cierre}.pdf"
    
    # Renderizar la plantilla en un string
    html_string = render_to_string(template_path, context)
    
    # Crear un objeto de BytesIO para generar el PDF
    pdf = io.BytesIO()
    
    # Crear el PDF
    pisa_status = pisa.CreatePDF(
        io.BytesIO(html_string.encode('utf-8')),
        dest=pdf
    )
    
    # Si no hubo errores, preparar la respuesta con el PDF
    if not pisa_status.err:
        pdf.seek(0)
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response
    
    # Si hubo un error, mostrar el error en la respuesta
    return HttpResponse(f"Error al generar el PDF: {pisa_status.err}", content_type='text/plain')







# Logger

from django.contrib.auth import views as auth_views
class CustomLoginView(auth_views.LoginView):
    template_name = 'login.html'  


#pagina para clinetes.
from django.shortcuts import render
from .models import Categoria, Menu, Mesa

def cliente(request):
    # Obtener todas las mesas
    mesas = Mesa.objects.all()

    # Obtener todas las categorías y organizar los ítems de menú por categoría, filtrando solo los disponibles
    categorias = Categoria.objects.all()
    menu_items = {categoria.nombre: Menu.objects.filter(categoria=categoria, disponible=True) for categoria in categorias}

    return render(request, 'restaurant/cliente.html', {
        'mesas': mesas,
        'categorias': categorias,
        'menu_items': menu_items,
    })



#Menu "Platos."
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .forms import MenuForm
from .models import Menu
from django.contrib.auth.decorators import login_required

@login_required
def listar_menu(request):
    # Agrupar los menús por categoría
    categorias = Menu.objects.values_list('categoria__nombre', flat=True).distinct()
    menu_categorias = {categoria: Menu.objects.filter(categoria__nombre=categoria, disponible=True) for categoria in categorias}
    menu_items_no_disponibles = Menu.objects.filter(disponible=False)

    return render(request, 'menu/listar_menu.html', {
        'menu_categorias': menu_categorias,
        'menu_items_no_disponibles': menu_items_no_disponibles
    })

from .models import Categoria
@login_required
def crear_menu(request):
    categorias = Categoria.objects.all()  # Obtener todas las categorías
    if request.method == 'POST':
        form = MenuForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'El menú ha sido guardado con éxito.')
            return redirect('listar_menu')
        else:
            messages.error(request, 'Error al guardar el menú. Verifica los datos ingresados.')
    else:
        form = MenuForm()

    return render(request, 'menu/crear_menu.html', {'form': form, 'categorias': categorias})

@login_required
def editar_menu(request, menu_id):
    menu_item = get_object_or_404(Menu, id=menu_id)
    categorias = Categoria.objects.all()  # Obtener todas las categorías
    
    if request.method == 'POST':
        form = MenuForm(request.POST, request.FILES, instance=menu_item)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Plato actualizado con éxito.')
            return redirect('listar_menu')
    else:
        form = MenuForm(instance=menu_item)
    
    return render(request, 'menu/editar_menu.html', {
        'form': form, 
        'menu': menu_item,
        'categorias': categorias  # Pasar las categorías al contexto
    })

@login_required
def eliminar_menu(request, menu_id):
    menu_item = get_object_or_404(Menu, id=menu_id)
    if request.method == 'POST':
        menu_item.delete()
        messages.success(request, 'Plato eliminado con éxito.')
        return redirect('listar_menu')
    return render(request, 'menu/eliminar_menu.html', {'menu': menu_item})

@login_required
def cambiar_disponibilidad_menu(request, menu_id):
    menu_item = get_object_or_404(Menu, id=menu_id)
    menu_item.disponible = not menu_item.disponible
    menu_item.save()
    messages.success(request, 'La disponibilidad del menú ha sido actualizada.')
    return redirect('listar_menu')

#crear categorias para menu
from .models import Categoria
from .forms import CategoriaForm

@login_required
def crear_categoria(request):
    if request.method == "POST":
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('listar_categorias')
        else:
            messages.error(request, 'Hubo un error al crear la categoría. Por favor, verifica los datos ingresados.')
    else:
        form = CategoriaForm()

    return render(request, 'menu/categoria/crear_categoria.html', {'form': form})

@login_required
def eliminar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    if request.method == "POST":
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente.')
        return redirect('listar_categorias')
    return render(request, 'menu/categoria/eliminar_categoria.html', {'categoria': categoria})

@login_required
def listar_categorias(request):
    categorias = Categoria.objects.all()
    return render(request, 'menu/categoria/listar_categorias.html', {'categorias': categorias})

from django.http import HttpResponse
from .models import Menu

def filtrar_platos_por_categoria(request):
    categoria_id = request.GET.get('categoria_id')
    platos = Menu.objects.filter(categoria_id=categoria_id, disponible=True) if categoria_id != '0' else Menu.objects.filter(disponible=True)
    html = ''
    for plato in platos:
        html += f'''
        <div class="card" style="width: 18rem; margin-right: 10px;">
            <img src="{plato.imagen.url}" class="card-img-top" alt="{plato.nombre_plato}">
            <div class="card-body">
                <h5 class="card-title">{plato.nombre_plato}</h5>
                <p class="card-text">${plato.precio}</p>
                <button class="btn btn-primary agregar-pedido" data-id="{plato.id}" data-precio="{plato.precio}">Agregar al Pedido</button>
            </div>
        </div>
        '''
    return HttpResponse(html)




#fin menu.


#pagos.

@login_required
def crear_pago(request):
    caja_abierta = Caja.objects.filter(estado='abierta').first()
    if not caja_abierta:
        messages.error(request, 'Debe abrir una caja antes de registrar pagos.')
        return redirect('apertura_caja')

    pedidos_pendientes = Pedido.objects.filter(estado='pendiente')

    if request.method == 'POST':
        form = PagoForm(request.POST)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.caja = caja_abierta
            pago.save()
            messages.success(request, 'Pago registrado exitosamente.')
            return redirect('consulta_caja', caja_id=caja_abierta.id)
    else:
        form = PagoForm()

    return render(request, 'caja/crear_pago.html', {'form': form, 'pedidos': pedidos_pendientes})

@login_required
def modificar_pago(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    if request.method == 'POST':
        form = PagoForm(request.POST, instance=pago)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pago modificado exitosamente.')
            return redirect('consulta_caja', caja_id=pago.caja.id)
    else:
        form = PagoForm(instance=pago)

    return render(request, 'caja/modificar_pago.html', {'form': form, 'pago': pago})

@login_required
def eliminar_pago(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    if request.method == 'POST':
        pago.delete()
        messages.success(request, 'Pago eliminado exitosamente.')
        return redirect('consulta_caja', caja_id=pago.caja.id)

    return render(request, 'caja/eliminar_pago.html', {'pago': pago})



from django.http import HttpResponseForbidden

@login_required
def mi_vista(request):
    if not request.user.has_perm('app_name.perm_name'):
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")
    
    
#Procesos que tienen excel.     

import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Proveedor
from django.contrib.auth.decorators import login_required
import openpyxl
from django.http import HttpResponse

@login_required
def descargar_plantilla_proveedores(request):
    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    # Definir los encabezados de las columnas según la estructura de la base de datos
    headers = ['Nombre', 'CUIT', 'CBU', 'Alias CBU', 'Calle', 'N°', 'Localidad', 'País', 'Código Postal', 'Teléfono', 'Email', 'Plazo de Pago', 'Observaciones']
    ws.append(headers)

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_proveedores.xlsx'

    wb.save(response)
    return response

@login_required
def cargar_proveedores_masivo(request):
    if request.method == "POST":
        if 'archivo_excel' in request.FILES:
            archivo_excel = request.FILES['archivo_excel']

            try:
                # Cargar el archivo Excel en un DataFrame de pandas
                df = pd.read_excel(archivo_excel)

                # Verificar encabezados correctos
                headers = ['Nombre', 'CUIT', 'CBU', 'Alias CBU', 'Calle', 'N°', 'Localidad', 'País', 'Código Postal', 'Teléfono', 'Email', 'Plazo de Pago', 'Observaciones']
                if list(df.columns) != headers:
                    mensajes = f"Encabezados incorrectos. Se esperaba: {headers} y se encontró {list(df.columns)}."
                    messages.error(request, mensajes)
                    return render(request, 'proveedores/cargar_proveedores.html')

                # Lista para almacenar los errores y advertencias
                errores = []
                advertencias = []

                # Iterar sobre las filas del DataFrame
                for index, row in df.iterrows():
                    columnas_vacias = []
                    if pd.isna(row['Nombre']):
                        columnas_vacias.append('Nombre')
                    if pd.isna(row['CUIT']):
                        columnas_vacias.append('CUIT')
                    if pd.isna(row['CBU']):
                        columnas_vacias.append('CBU')
                    if pd.isna(row['Alias CBU']):
                        columnas_vacias.append('Alias CBU')
                    if pd.isna(row['Calle']):
                        columnas_vacias.append('Calle')
                    if pd.isna(row['N°']):
                        columnas_vacias.append('N°')
                    if pd.isna(row['Localidad']):
                        columnas_vacias.append('Localidad')
                    if pd.isna(row['País']):
                        columnas_vacias.append('País')
                    if pd.isna(row['Código Postal']):
                        columnas_vacias.append('Código Postal')
                    if pd.isna(row['Teléfono']):
                        columnas_vacias.append('Teléfono')
                    if pd.isna(row['Email']):
                        columnas_vacias.append('Email')
                    if pd.isna(row['Plazo de Pago']):
                        columnas_vacias.append('Plazo de Pago')
                    if pd.isna(row['Observaciones']):
                        columnas_vacias.append('Observaciones')

                    if columnas_vacias:
                        advertencias.append(f"Fila {index + 1}: Las siguientes columnas están vacías: {', '.join(columnas_vacias)}")

                    try:
                        Proveedor.objects.create(
                            nombre_proveedor=row['Nombre'],
                            cuit=row.get('CUIT', ''),
                            cbu=row.get('CBU', ''),
                            alias_cbu=row.get('Alias CBU', ''),
                            calle=row.get('Calle', ''),
                            numero=row.get('N°', ''),
                            localidad=row.get('Localidad', ''),
                            pais=row.get('País', ''),
                            codigo_postal=row.get('Código Postal', ''),
                            telefono=row.get('Teléfono', ''),
                            email=row.get('Email', ''),
                            plazo_pago=row.get('Plazo de Pago', None),
                            observaciones=row.get('Observaciones', '')
                        )
                    except Exception as e:
                        errores.append(f"Fila {index + 1}: {str(e)}")

                if errores:
                    messages.error(request, "Se encontraron errores al cargar algunos proveedores.")
                    return render(request, 'proveedores/cargar_proveedores.html', {'errores': errores})
                else:
                    messages.success(request, "Proveedores cargados exitosamente.")
                    if advertencias:
                        messages.warning(request, "Advertencias encontradas en el archivo:")
                    return render(request, 'proveedores/cargar_proveedores.html', {'advertencias': advertencias})

            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return redirect('cargar_proveedores_masivo')
        else:
            messages.error(request, "Por favor, sube un archivo Excel.")
            return redirect('cargar_proveedores_masivo')

    return render(request, 'proveedores/cargar_proveedores.html')


#para inventario 

from .models import Inventario
@login_required
def descargar_plantilla_inventario(request):
    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Definir los encabezados de las columnas
    headers = ['Nombre Producto', 'Cantidad Actual', 'Cantidad Mínima', 'Unidad de Medida']
    ws.append(headers)

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_inventario.xlsx'

    wb.save(response)
    return response

@login_required
def cargar_inventario_masivo(request):
    if request.method == "POST":
        if 'archivo_excel' in request.FILES:
            archivo_excel = request.FILES['archivo_excel']

            try:
                # Cargar el archivo Excel en un DataFrame de pandas
                df = pd.read_excel(archivo_excel)

                # Verificar encabezados correctos
                headers = ['Nombre Producto', 'Cantidad Actual', 'Cantidad Mínima', 'Unidad de Medida']
                if list(df.columns) != headers:
                    mensajes = f"Encabezados incorrectos. Se esperaba: {headers} y se encontró {list(df.columns)}."
                    messages.error(request, mensajes)
                    return render(request, 'inventario/excel/cargar_inventario.html')

                # Lista para almacenar los errores y advertencias
                errores = []
                advertencias = []

                # Iterar sobre las filas del DataFrame
                for index, row in df.iterrows():
                    columnas_vacias = []
                    if pd.isna(row['Nombre Producto']):
                        columnas_vacias.append('Nombre Producto')
                    if pd.isna(row['Cantidad Actual']):
                        columnas_vacias.append('Cantidad Actual')
                    if pd.isna(row['Cantidad Mínima']):
                        columnas_vacias.append('Cantidad Mínima')
                    if pd.isna(row['Unidad de Medida']):
                        columnas_vacias.append('Unidad de Medida')

                    if columnas_vacias:
                        advertencias.append(f"Fila {index + 1}: Las siguientes columnas están vacías: {', '.join(columnas_vacias)}")

                    try:
                        Inventario.objects.update_or_create(
                            nombre_producto=row['Nombre Producto'],
                            defaults={
                                'cantidad_actual': row['Cantidad Actual'],
                                'cantidad_minima': row['Cantidad Mínima'],
                                'unidad_medida': row['Unidad de Medida']
                            }
                        )
                    except Exception as e:
                        errores.append(f"Fila {index + 1}: {str(e)}")

                if errores:
                    messages.error(request, "Se encontraron errores al cargar algunos productos.")
                    return render(request, 'inventario/excel/cargar_inventario.html', {'errores': errores})
                else:
                    messages.success(request, "Inventario cargado exitosamente.")
                    if advertencias:
                        messages.warning(request, "Advertencias encontradas en el archivo:")
                        return render(request, 'inventario/excel/cargar_inventario.html', {'advertencias': advertencias})

            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return redirect('cargar_inventario_masivo')
        else:
            messages.error(request, "Por favor, sube un archivo Excel.")
            return redirect('cargar_inventario_masivo')

    return render(request, 'inventario/excel/cargar_inventario.html')

@login_required
def exportar_inventario(request):
    # Exportar el inventario actual
    inventario = Inventario.objects.all()

    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Actual"

    # Definir los encabezados de las columnas
    headers = ['Nombre Producto', 'Cantidad Actual', 'Cantidad Mínima', 'Unidad de Medida']
    ws.append(headers)

    # Rellenar con los datos actuales del inventario
    for item in inventario:
        ws.append([
            item.nombre_producto,
            item.cantidad_actual,
            item.cantidad_minima,
            item.unidad_medida
        ])

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventario_actual.xlsx'

    wb.save(response)
    return response


#para la cocina 
from django.shortcuts import render
from .models import Mesa, Pedido
from datetime import datetime
from django.utils import timezone

def cocinero_dashboard(request):
    # Obtener todos los pedidos en estado "preparando"
    pedidos = Pedido.objects.filter(estado='preparando').order_by('fecha_pedido')

    # Para cada pedido, obtener los detalles del pedido
    for pedido in pedidos:
        pedido.detalles_pedido = DetallePedido.objects.filter(pedido=pedido)
        # Calcular el tiempo transcurrido en minutos
        pedido.tiempo = (timezone.now() - pedido.fecha_pedido).total_seconds() / 60

    context = {
        'pedidos': pedidos,
    }
    return render(request, 'cocina/cocineros.html', context)



#Registrar pago.

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Pedido, Pago, Caja, TransaccionCaja
from .forms import PagoForm
from django.utils import timezone

@login_required
def registrar_pago(request, pedido_id):
    caja_abierta = Caja.objects.filter(estado='abierta', usuario=request.user).first()
    if not caja_abierta:
        return redirect('apertura_caja')
    pedido = get_object_or_404(Pedido, id=pedido_id, estado__in=['servido', 'pendiente'])
    if request.method == 'POST':
        metodo_pago = request.POST.get('metodo_pago')
        monto_pagado = request.POST.get('monto_pagado')
        if metodo_pago == 'efectivo':
            try:
                
                monto_pagado = float(monto_pagado)
            except (ValueError, TypeError):
                messages.error(request, 'El monto pagado no es válido. Por favor ingrese un número.')
                return redirect('registrar_pago', pedido_id=pedido.id)

            if monto_pagado < pedido.total:
                messages.error(request, 'El monto pagado no puede ser menor que el total del pedido.')
                return redirect('registrar_pago', pedido_id=pedido.id)
        else:
            monto_pagado = pedido.total  
        pago = Pago(pedido=pedido, metodo_pago=metodo_pago, monto=monto_pagado)
        pago.save()
        pedido.estado = 'pagado'
        pedido.save()
        transaccion = TransaccionCaja(
            caja=caja_abierta,
            tipo='ingreso',
            monto=pago.monto,
            descripcion=f'Pago de pedido {pedido.id}',
            pago=pago  
        )
        transaccion.save()

        messages.success(request, 'Pago registrado con éxito.')
        return redirect('pedidos_activos')

    return render(request, 'caja/registrar_pago.html', {'pedido': pedido, 'caja': caja_abierta})


#estadisticas

from django.shortcuts import render
from .forms import ReporteFiltroForm
from .models import Caja, Pedido, TransaccionCaja, Usuario
from django.db.models import Sum

def generar_reporte(request):
    form = ReporteFiltroForm(request.POST or None)
    reportes = []
    
    if form.is_valid():
        fecha_inicio = form.cleaned_data['fecha_inicio']
        fecha_fin = form.cleaned_data['fecha_fin']
        usuario = form.cleaned_data['usuario']
        
        # Filtro de cajas abiertas o cerradas
        cajas = Caja.objects.filter(apertura__date__gte=fecha_inicio, apertura__date__lte=fecha_fin)
        
        if usuario:
            cajas = cajas.filter(usuario=usuario)
        
        for caja in cajas:
            transacciones = caja.transacciones.all()
            pedidos = Pedido.objects.filter(fecha_pedido__gte=caja.apertura, fecha_pedido__lte=caja.cierre or timezone.now())

            total_ingresos = transacciones.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
            total_egresos = transacciones.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or 0
            
            reportes.append({
                'caja': caja,
                'total_ingresos': total_ingresos,
                'total_egresos': total_egresos,
                'pedidos': pedidos,
            })
    
    return render(request, 'reportes/generar_reporte.html', {
        'form': form,
        'reportes': reportes
    })


import plotly.express as px
from django.db.models import Sum
from .models import TransaccionCaja

def obtener_datos_grafico_ingresos(request):
    # Obtener los ingresos por día
    ingresos_por_dia = TransaccionCaja.objects.filter(tipo='ingreso').values('fecha').annotate(total_ingresos=Sum('monto'))

    # Extraer los datos
    fechas = [ingreso['fecha'].strftime('%Y-%m-%d') for ingreso in ingresos_por_dia]
    total_ingresos = [ingreso['total_ingresos'] for ingreso in ingresos_por_dia]

    # Crear gráfico con Plotly
    fig = px.line(x=fechas, y=total_ingresos, labels={'x': 'Fecha', 'y': 'Ingresos'}, title='Ingresos Totales por Día')

    # Convertir el gráfico a JSON
    graph_json = fig.to_json()

    return JsonResponse(graph_json, safe=False)

def obtener_datos_grafico_metodos_pago(request):
    # Obtener los totales por método de pago
    metodos_pago_totales = TransaccionCaja.objects.filter(tipo='ingreso').exclude(pago__metodo_pago__isnull=True).values('pago__metodo_pago').annotate(total=Sum('monto'))


    # Extraer los datos
    metodos_pago = [metodo['pago__metodo_pago'] for metodo in metodos_pago_totales]
    totales = [metodo['total'] for metodo in metodos_pago_totales]

    # Crear gráfico con Plotly
    fig = px.pie(values=totales, names=metodos_pago, title='Métodos de Pago')

    # Convertir el gráfico a JSON
    graph_json = fig.to_json()

    return JsonResponse(graph_json, safe=False)





from django.contrib import messages
from .models import Usuario
from .forms import CustomUserCreationForm, CustomUserChangeForm
from .models import Usuario


@login_required
def crear_usuario(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('listar_usuarios')
        else:
            messages.error(request, 'Error al crear el usuario. Verifique los datos ingresados.')
    else:
        form = CustomUserCreationForm()
    return render(request, 'usuarios/crear_usuario.html', {'form': form})

def listar_usuarios(request):
    usuarios = Usuario.objects.all()
    return render(request, 'usuarios/listar_usuarios.html', {'usuarios': usuarios})


@login_required
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario editado exitosamente.')
            return redirect('listar_usuarios')
    else:
        form = CustomUserChangeForm(instance=usuario)
    return render(request, 'usuarios/editar_usuario.html', {'form': form})

@login_required
def eliminar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    if request.method == 'POST':
        usuario.delete()
        messages.success(request, 'Usuario eliminado exitosamente.')
        return redirect('listar_usuarios')
    return render(request, 'usuarios/eliminar_usuario.html', {'usuario': usuario})



from django.contrib.auth import update_session_auth_hash
from .models import Usuario
from .forms import CustomPasswordResetForm

@login_required
def restablecer_contraseña(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    if request.method == 'POST':
        form = CustomPasswordResetForm(request.POST)
        if form.is_valid():
            nueva_contraseña = form.cleaned_data['nueva_contraseña']
            usuario.set_password(nueva_contraseña)
            usuario.save()
            update_session_auth_hash(request, usuario)  # Mantener la sesión activa para el administrador si es necesario
            messages.success(request, 'Contraseña restablecida exitosamente.')
            return redirect('listar_usuarios')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = CustomPasswordResetForm()

    return render(request, 'usuarios/restablecer_contraseña.html', {'form': form, 'usuario': usuario})

from .forms import CustomUserCreationForm
from .models import Usuario

def registrar_usuario(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            # Crear usuario como inactivo
            usuario = form.save(commit=False)
            usuario.is_active = False  # El administrador debe activarlo
            usuario.save()
            messages.success(request, 'Tu cuenta ha sido creada. Un administrador la activará pronto.')
            return redirect('login')
        else:
            messages.error(request, 'Error al registrar el usuario. Verifica los datos ingresados.')
    else:
        form = CustomUserCreationForm()

    return render(request, 'usuarios/registrar_usuario.html', {'form': form})


from .forms import CustomPasswordResetForm

def recuperar_contraseña(request):
    mostrar_nueva_contraseña = False  # Controla si se muestran los campos de nueva contraseña

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        
        # Intentar encontrar al usuario con el username y correo proporcionados
        try:
            usuario = Usuario.objects.get(username=username, email=email)
            
            # Si ya se pasó la validación, mostrar el formulario de nueva contraseña
            form = CustomPasswordResetForm(request.POST)
            if 'nueva_contraseña' in request.POST and form.is_valid():
                nueva_contraseña = form.cleaned_data['nueva_contraseña']
                usuario.set_password(nueva_contraseña)
                usuario.save()
                
                messages.success(request, 'Tu contraseña ha sido restablecida exitosamente.')
                return redirect('login')
            else:
                mostrar_nueva_contraseña = True
                messages.info(request, 'Por favor ingresa una nueva contraseña.')
                
        except Usuario.DoesNotExist:
            messages.error(request, 'Los datos ingresados no son correctos.')
            form = CustomPasswordResetForm()  # Formulario vacío para nueva contraseña
    else:
        form = CustomPasswordResetForm()

    return render(request, 'usuarios/recuperar_contraseña.html', {'form': form, 'mostrar_nueva_contraseña': mostrar_nueva_contraseña})
